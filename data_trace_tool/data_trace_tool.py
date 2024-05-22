from PyQt5.QtGui import *
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QTableWidgetItem, QMessageBox, QComboBox, QTableWidget
from PyQt5.QtCore import *
from PyQt5 import QtWidgets, QtGui
import requests
import pymysql
from data_trace_tool_ui import Ui_MainWindow
from qss import qss
from threading import Thread
from pymysql.err import OperationalError
import openpyxl


# 拖拽事件参考https://blog.csdn.net/hubing_hust/article/details/128072839

class DataTraceTool(QMainWindow, Ui_MainWindow):
    signal = pyqtSignal(bool,str)
    # updateTableWidgetSignal = pyqtSignal(str)

    def __init__(self, parent=None):
        super(DataTraceTool, self).__init__(parent)
        self.setupUi(self)
        self.initUI()
        self.user_id = None
        self.prevRow = -1
        self.action_key_num = 0
        self.page_key_num = 0

        self.comboBox.currentIndexChanged.connect(self.handleSelectionChange)
        self.uploadFileButton.clicked.connect(self.openfile)
        # self.uploadFileButton.clicked.connect(self.creat_table_show)
        self.tableWidget.itemClicked.connect(self.show_sql)
        self.getUserIdButton.clicked.connect(self.get_user_id)
        self.testConnectPushButton.clicked.connect(self.test_database_connect)
        self.SQLPushButton.clicked.connect(self.get_result)
        self.signal.connect(self.show_message)
        # self.updateTableWidgetSignal.connect(self.creat_table_show)

    def initUI(self):
        # 充值窗口大小为屏幕的比例
        # 获取显示器分辨率
        self.screenRect = QtWidgets.QApplication.desktop().screenGeometry()
        self.resize(3 * self.screenRect.width() / 5, 3 * self.screenRect.height() / 5)
        self.getUserIdButton.setEnabled(False)
        # 创建tableWidget，支持拖拽上传
        self.tableWidget = MyTableWidget(self.groupBox_2)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.verticalLayout.addWidget(self.tableWidget)
        # 初始化部分值
        self.ipLineEdit.setText("bj-cdb-4kqb4cis.sql.tencentcdb.com")
        self.portLineEdit.setText("61292")
        self.dataBaseNameLineEdit.setText("data_trace")
        self.userNameLineEdit.setText("xd_dev_mysql")
        self.dataBasePasswordLineEdit.setText("Ycc3lgu1rjc4q8qe")
        self.tableWidget.updateTableSignal.connect(self.updateTableWidget)

    def show_message(self, flag=None, message=None):
        if flag:
            QMessageBox.information(None, "成功", message)
        else:
            QMessageBox.critical(None, "失败", message)

    def handleSelectionChange(self):
        self.userIdLineEdit.clear()
        self.user_id = None
        # self.passwordLineEdit.clear()
        if self.comboBox.currentText() == 'APP用户请输入user_id':
            self.getUserIdButton.setEnabled(False)
        else:
            self.getUserIdButton.setEnabled(True)

    def openfile(self):
        # 获取路径
        openfile_name = QFileDialog.getOpenFileName(None,'选择文件','','Excel files(*.xlsx , *.xls)')
        # 获取路径
        path_openfile_name = openfile_name[0]
        if len(path_openfile_name) > 0:
            self.updateTableWidget(path_openfile_name)

    def show_sql(self, Item=None):
        self.action_key_num = self.actionKeyLineEdit.text()
        self.page_key_num = self.pageKeyLineEdit.text()
        if self.comboBox.currentText() == 'APP用户请输入user_id':
            self.user_id = self.userIdLineEdit.text().strip()
        if (self.user_id is None) or (len(self.user_id)==0):
            QMessageBox.critical(None, '错误', '请先获取user_id')
            return
        else:
            if Item is None:
                return
            else:
                if self.prevRow != -1:
                    for i in range(1, self.tableWidget.columnCount()):
                        self.tableWidget.item(self.prevRow,i).setBackground(QColor(255,255,255))
                row = Item.row()
                for i in range(1, self.tableWidget.columnCount()):
                    self.tableWidget.item(row,i).setBackground(QColor(240,255,255))
                self.sqlShowTextEdit.setText("SELECT * FROM data_trace.t_action_info WHERE  1 "
                                                    +"\n  AND user_id = '"+self.user_id
                                                    +"'\n  AND page_key = '"+self.tableWidget.item(row, int(self.page_key_num)).text()
                                                    +"'\n  AND action_key = '"+self.tableWidget.item(row, int(self.action_key_num)).text()
                                                    +"'\n  AND DATE_FORMAT(action_time,'%y-%M-%d') = DATE_FORMAT(now(),'%y-%M-%d') "
                                                    +"\nORDER BY action_time DESC limit 10;")
                self.prevRow = row
                self.get_result()

    def get_user_id(self):
        choice = self.comboBox.currentText().strip()
        user263 = self.userIdLineEdit.text().strip()
        web_header = {
            "Content-Type": "application/json;charset=UTF-8"            
        }
        web_url = "https://test_model-cloud-bf.dayustudy.com/inner/inner/getUserInfo"
        web_param = {
            "userName": user263,
            "systemType": 2
        }
        if len(user263) == 0:
            QMessageBox.critical(None, '错误', '请输入263点击获取user_id或直接输入APP中的user_id')
            return
        else:
            if choice=="大鱼或企微用户263":
                try:
                    resp = requests.get(url=web_url, params=web_param, headers=web_header)
                except Exception as e:
                    QMessageBox.critical(None, '请求异常', str(e))
                else:
                    self.user_id = str(resp.json().get("data").get("userId"))
            elif choice=="APP用户手机号":
                self.user_id = user263
            if self.user_id is not None:
                QMessageBox.information(None, '获取成功', 'user_id为'+self.user_id)
            else:
                QMessageBox.critical( None, '错误','user_id获取失败！')
                return

    def test_database_connect(self):
        ip = self.ipLineEdit.text()
        port = int(self.portLineEdit.text().strip())
        database_name = self.dataBaseNameLineEdit.text()
        username = self.userNameLineEdit.text()
        password = self.dataBasePasswordLineEdit.text()
        # 创建新的线程去执行测试链接数据库方法，
        # 服务器慢，只会在新线程中阻塞
        # 不影响主线程
        thread = Thread(target=self.test_connect_thread, args=(ip, username, password, database_name, port))
        thread.start()

    def test_connect_thread(self, ip, username, password, database_name, port):
        try:
            conn = pymysql.connect(host=ip, user=username, password=password, database=database_name, port=port, connect_timeout=10)
            self.signal.emit(True, "数据库连接成功")
            conn.close()
            return
        except Exception as e:
            self.signal.emit(False, str(e))
            return

    def get_result(self):
        ip = self.ipLineEdit.text().strip()
        port = int(self.portLineEdit.text().strip())
        database_name = self.dataBaseNameLineEdit.text().strip()
        username = self.userNameLineEdit.text().strip()
        password = self.dataBasePasswordLineEdit.text().strip()
        sql = self.sqlShowTextEdit.toPlainText().strip()
        thread = Thread(target=self.get_result_thread, args=(ip, username, password, database_name, port, sql))
        thread.start()

    def get_result_thread(self, ip, username, password, database_name, port, sql):
        try:
            conn = pymysql.connect(host=ip, user=username, password=password, database=database_name, port=port, read_timeout=5)
            cursor = conn.cursor()
            cursor.execute(sql)
            data = cursor.fetchall()
            desc = cursor.description
            table_header = [item[0] for item in desc]
            self.statusbar.showMessage("共查询到%d条埋点数据" %(len(data)))
            model = QStandardItemModel(len(data), len(desc))
            model.setHorizontalHeaderLabels(table_header)
            for column_num in range(len(desc)):
                head_item = model.horizontalHeaderItem(column_num)
                head_item.setFont(QFont("宋体",12,QFont.Bold))
                head_item.setBackground(QBrush(QColor(192, 192, 192)))
                # 设置字体颜色为黑色
                head_item.setForeground(QBrush(QColor(255,255,255)))
            for row in range(len(data)):
                for column in range(len(desc)):
                    item = QStandardItem(str(data[row][column]))
                    model.setItem(row, column, item)
            self.tableView.setModel(model)
            conn.close()
            return
        except AttributeError as ae:
            self.signal.emit(False,str(ae))
        except OperationalError as oe:
            # print(str(oe))
            self.signal.emit(False, '查询超时')
        except Exception as e:
            self.signal.emit(False, str(e))
            return

    def updateTableWidget(self, file_path):
        # 读取表格，转换表格
        if len(file_path) > 0:
            wb =  openpyxl.load_workbook(file_path)
            ws = wb.active
            # 获取行数和列数
            rows = ws.max_row
            cols = ws.max_column
            # tablewidget设置行数和列数
            self.tableWidget.setRowCount(rows-1)    # 注意行数要减掉表头
            self.tableWidget.setColumnCount(cols+1) # 注意列数要加上左侧增加的1列
            # 获取表头数据
            table_data = []
            for row in ws.iter_rows(min_row=1, max_col=cols, max_row=rows, values_only=True):
                table_data.append(list(row))
            table_head_data = table_data[0]
            table_head_data.insert(0, "测试结果")
            # 获取上传表格中的page_key和action_key的列数
            self.page_key_num = table_head_data.index("page_key")
            self.action_key_num = table_head_data.index("action_key")
            self.tableWidget.setHorizontalHeaderLabels(table_head_data)
            for column_num in range(cols+1):
                head_item = self.tableWidget.horizontalHeaderItem(column_num)
                head_item.setFont(QFont("宋体", 12, QFont.Bold))
                head_item.setBackground(QBrush(QColor(192, 192, 192)))
                # 设置字体颜色为白色
                head_item.setForeground(QBrush(QColor(255, 255, 255)))
            # 遍历表格每个元素，同时添加到tablewidget中
            print(rows)
            for i in range(0, rows-1):
                comBox = QComboBox()
                comBox.addItems(['成功', '失败'])
                comBox.setCurrentIndex(-1)
                self.tableWidget.setCellWidget(i, 0, comBox)
                for j in range(0, cols):
                    # 将遍历的元素添加到tablewidget中并显示
                    newItem = QTableWidgetItem('' if table_data[i+1][j] is None else str(table_data[i+1][j]))
                    newItem.setFlags(Qt.ItemIsEnabled)
                    newItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.tableWidget.setItem(i, j+1, newItem)


class MyTableWidget(QTableWidget):
    updateTableSignal = pyqtSignal(str)
    def __init__(self, parent=None):
        super(MyTableWidget, self).__init__(parent)
        self.setAcceptDrops(True)
    # 设置拖拽文件到该表格，触发该方法
    def dragEnterEvent(self, event):
        self.file_path = event.mimeData().text()
        if '.xls' in self.file_path or ('.xlsx' in self.file_path):
            event.accept()
            # print(self.file_path.replace('file:///', '').replace("/", "\\"))
            self.file_path = self.file_path.replace('file:///', '').replace("/", "\\")
        else:
            event.ignore()
            # print('文件非excel文件')

    def dragMoveEvent(self, event):
        # print("dragMoveEvent")
        event.accept()

    # 注意此方法必须在dragMoveEvent接受后触发
    # 用户释放鼠标按钮时，这个方法会被触发，用于处理拖拽事件
    def dropEvent(self, event):
        # print('释放后触发')
        # self.updateTableWidget(self.file_path)
        self.updateTableSignal.emit(self.file_path)



if __name__ == "__main__":
    app = QApplication([])
    app.setWindowIcon(QIcon("tool.ico"))
    # 读取样式表
    # with open("data_trace_tool.qss", "r") as f:
    #     app.setStyleSheet(f.read())
    app.setStyleSheet(qss)
    dataTraceTool = DataTraceTool()
    dataTraceTool.show()
    # 设置程序运行样式为融合风格，不然表头背景色不会生效
    app.setStyle(QtWidgets.QStyleFactory.create('fusion'))
    app.exec_()
