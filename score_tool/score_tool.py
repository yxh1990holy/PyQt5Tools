from PyQt5.QtGui import *
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QTableWidgetItem, QMessageBox, QComboBox
from PyQt5.QtCore import *
from PyQt5 import QtWidgets, QtCore
from score_ui import Ui_MainWindow
import openpyxl

class ScoreTool(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.main_data = []
        self.student_num = {}
        self.average_score = {}
        self.setupUi(self)
        self.screenRect = QtWidgets.QApplication.desktop().screenGeometry()
        self.resize(3 * self.screenRect.width() / 5, 3 * self.screenRect.height() / 5)

        self.uploadFileButton.clicked.connect(self.upload_file)
        self.caculateAverageScoreButton.clicked.connect(self.get_average)
        self.saveToExcelButton.clicked.connect(self.save_to_excel)

    # 上传文件并显示对应表数据
    def upload_file(self):

        # 获取路径
        openfile_name = QFileDialog.getOpenFileName(None, '选择文件', '', 'Excel files(*.xlsx , *.xls)')
        # 获取路径
        path_openfile_name = openfile_name[0]
        if len(path_openfile_name) == 0:
            return 
        # 获取表中所有sheet
        wb = openpyxl.load_workbook(path_openfile_name)
        sheetnames_list = wb.sheetnames
        for sheet_name in sheetnames_list:
            tab = QtWidgets.QWidget()
            tab.setObjectName(sheet_name)
            vertical_layout_current = QtWidgets.QVBoxLayout(tab)
            vertical_layout_current.setObjectName("vertical_layout_current")
            table_widget = QtWidgets.QTableWidget(tab)
            width = self.tabWidget.geometry().width()
            height = self.tabWidget.geometry().height()
            table_widget.setGeometry(QtCore.QRect(0, 0, width, height-39))
            table_widget.setObjectName("table_widget")
            ws = wb[sheet_name]
            max_column = ws.max_column
            max_row = self.get_max_row(ws)
            table_widget.setColumnCount(max_column)
            table_widget.setRowCount(max_row-2)

            # 获取数据并设置数据
            data = []
            for row in ws.iter_rows(min_row=2, max_col=max_column, max_row=max_row, values_only=True):
                data.append(list(row))
            # 设置表头数据
            table_widget.setHorizontalHeaderLabels(data[0])
            # 设置表格数据
            for i in range(1, max_row-1):
                for j in range(0, max_column):
                    # 将遍历的元素添加到table_widget中并显示
                    newItem = QTableWidgetItem(str(data[i][j]))
                    newItem.setFlags(Qt.ItemIsEnabled)
                    newItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    table_widget.setItem(i-1, j, newItem)
            # 设置表头样式
            for column_num in range(max_column):
                head_item = table_widget.horizontalHeaderItem(column_num)
                head_item.setFont(QFont("宋体", 12, QFont.Bold))
                head_item.setBackground(QBrush(QColor(192, 192, 192)))
                # 设置字体颜色为白色
                head_item.setForeground(QBrush(QColor(255, 255, 255)))
            vertical_layout_current.addWidget(table_widget)
            self.tabWidget.addTab(tab, sheet_name)
            # 存储总表的数据
            if sheet_name == "总体":
                self.main_data = data
            else:
                self.student_num[sheet_name] = max_row - 2
        wb.close()

   # 处理数据并显示平均成绩
    def get_average(self):
        result = {}
        if len(self.main_data)==0 :
            return
        for i in range(1, len(self.main_data)):
            if self.main_data[i][2] not in result.keys():
                result[self.main_data[i][2]] = [self.main_data[i][3],self.main_data[i][6]]+self.main_data[i][7:26:2]+self.main_data[i][27:36]
            else:
                result[self.main_data[i][2]] = [result[self.main_data[i][2]][0]+self.main_data[i][3],result[self.main_data[i][2]][1]+self.main_data[i][6],result[self.main_data[i][2]][2]+self.main_data[i][7],
                                                result[self.main_data[i][2]][3]+self.main_data[i][9],result[self.main_data[i][2]][4]+self.main_data[i][11],result[self.main_data[i][2]][5]+self.main_data[i][13],
                                                result[self.main_data[i][2]][6]+self.main_data[i][15],result[self.main_data[i][2]][7]+self.main_data[i][17],result[self.main_data[i][2]][8]+self.main_data[i][19],
                                                result[self.main_data[i][2]][9]+self.main_data[i][21],result[self.main_data[i][2]][10]+self.main_data[i][23],result[self.main_data[i][2]][11]+self.main_data[i][25],
                                                result[self.main_data[i][2]][12]+self.main_data[i][27],result[self.main_data[i][2]][13]+self.main_data[i][28],result[self.main_data[i][2]][14]+self.main_data[i][29],
                                                result[self.main_data[i][2]][15]+self.main_data[i][30],result[self.main_data[i][2]][16]+self.main_data[i][31],result[self.main_data[i][2]][17]+self.main_data[i][32],
                                                result[self.main_data[i][2]][18]+self.main_data[i][33],result[self.main_data[i][2]][19]+self.main_data[i][34],result[self.main_data[i][2]][20]+self.main_data[i][35]]
        for key,student_score_list in result.items():
            result[key] = ['{:.2f}'.format(item/self.student_num[key]) for item in student_score_list]
        result = dict(sorted(result.items()))
        # 获取表头数据
        avg_score_table_header = ['班级',self.main_data[0][3]+'平均分',self.main_data[0][6]+'平均分']+[str(item)+'题平均分' for item in self.main_data[0][7:26:2]]+[str(item)+'题平均分' for item in self.main_data[0][27:36]]
        # print(avg_score_table_header)
        self.tableWidget_2.setColumnCount(len(avg_score_table_header))
        self.tableWidget_2.setRowCount(len(result))
        # 设置表头数据
        self.tableWidget_2.setHorizontalHeaderLabels(avg_score_table_header)
        # 设置表格数据
        row_num = 0
        for class_name,score_list in result.items():
            newItem = QTableWidgetItem(str(class_name))
            newItem.setFlags(Qt.ItemIsEnabled)
            newItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            self.tableWidget_2.setItem(row_num, 0, newItem)
            # print(score_list)
            for i, score in enumerate(score_list):
                # 将遍历的元素添加到tablewidget中并显示
                newItem = QTableWidgetItem(str(score))
                newItem.setFlags(Qt.ItemIsEnabled)
                newItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                self.tableWidget_2.setItem(row_num, i+1, newItem)
            row_num += 1
        self.average_score = result
        self.average_score['表头'] = avg_score_table_header
        print(self.average_score)


    # 保存平均分到excel
    def save_to_excel(self):
        file_path, type = QFileDialog.getSaveFileName(self, '文件保存', '/', 'xlsx(*.xlsx)')
        if len(file_path) == 0:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '平均分'
        for i, head in enumerate(self.average_score['表头']):
            ws.cell(row=1, column=i+1, value=head)
        row_num = 2
        self.average_score.pop('表头')
        for class_name, score_list in self.average_score.items():
            ws.cell(row=row_num, column=1, value=class_name)
            for i, score in enumerate(score_list):
                ws.cell(row=row_num, column=i + 2, value=score)
            row_num += 1
        wb.save(file_path)
        print("写入完毕")


    # 获取excel最大行数
    def get_max_row(self, sheet):
        i =  sheet.max_row
        real_max_row = 0
        while i >0:
            row_dict = {i.value for i in sheet[i]}
            if row_dict == {None}:
                i = i-1
            else:
                real_max_row = i
                break
        return real_max_row



if __name__ == "__main__":
    app = QApplication([])
    app.setWindowIcon(QIcon("tool.ico"))
    # 读取样式表
    # with open("data_trace_tool.qss", "r") as f:
    #     app.setStyleSheet(f.read())
    # app.setStyleSheet(qss)
    scoreTool = ScoreTool()
    scoreTool.show()
    # 设置程序运行样式为融合风格，不然表头背景色不会生效
    app.setStyle(QtWidgets.QStyleFactory.create('fusion'))
    app.exec_()